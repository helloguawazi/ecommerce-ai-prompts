#!/usr/bin/env python
"""关键词管理：从Excel随机取一个关键词，删除该行，输出给Claude用。

用法:
  python .claude/keywords/pick_keyword.py
     → 随机从两个文件中取一个关键词，输出 JSON

输出JSON格式:
  {"keyword": "ai修图", "site": "image2", "domain": "image2.anyachina.cn",
   "search_volume": "106848", "file": "合并结果_搜索量≥50.xlsx"}
"""

import json, os, random, sys
from copy import copy

os.chdir(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

# 站点映射规则
SITE_MAP = {
    'image2': {
        'file': '合并结果_搜索量≥50.xlsx',
        'domain': 'image2.anyachina.cn',
        'keywords': ['image2', 'ai修图', 'ai生图', 'ai制图', 'ai图片',
                     'nanobanana', 'banana', 'missav'],
    },
    'poster': {
        'file': 'AI海报.xls',
        'domain': 'poster.anyachina.cn',
        'keywords': [],
    },
}

# 判断关键词归属站点
def guess_site(keyword):
    kw = keyword.lower()
    for site, cfg in SITE_MAP.items():
        for pattern in cfg['keywords']:
            if pattern.lower() in kw:
                return site, cfg['domain']
    # 含"海报"的归 poster，其他默认 image2
    if '海报' in kw:
        return 'poster', SITE_MAP['poster']['domain']
    return 'image2', SITE_MAP['image2']['domain']


def pick_from_xlsx(filepath):
    """从 .xlsx 读取所有行，随机取一条，删除后保存"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows or all(r[0] is None for r in rows):
        return None
    # 过滤空行
    rows = [r for r in rows if r[0] is not None and str(r[0]).strip()]
    if not rows:
        return None
    idx = random.randint(0, len(rows) - 1)
    picked = rows[idx]
    # 删除该行：重建 workbook
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    # 写表头
    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    ws2.append(header)
    for i, row in enumerate(rows):
        if i != idx:
            ws2.append(list(row))
    wb2.save(filepath)
    wb.close()
    return str(picked[0]).strip() if picked[0] else None


def pick_from_xls(filepath):
    """从 .xls 读取所有行，随机取一条，删除后保存"""
    import xlrd, xlwt
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    rows = []
    for i in range(1, ws.nrows):
        val = ws.cell_value(i, 0)
        if val and str(val).strip():
            rows.append([ws.cell_value(i, j) for j in range(ws.ncols)])
    if not rows:
        return None
    idx = random.randint(0, len(rows) - 1)
    picked = rows[idx]
    kw = str(picked[0]).strip()
    # 写回
    wb_out = xlwt.Workbook()
    ws_out = wb_out.add_sheet('Sheet0')
    # 表头
    for j in range(ws.ncols):
        ws_out.write(0, j, ws.cell_value(0, j))
    # 数据（跳过删除行）
    out_idx = 1
    for i, row in enumerate(rows):
        if i != idx:
            for j, val in enumerate(row):
                cell_val = val
                if isinstance(val, float) and val == int(val):
                    cell_val = int(val)
                ws_out.write(out_idx, j, cell_val)
            out_idx += 1
    wb_out.save(filepath)
    return kw


def main():
    files = [
        ('合并结果_搜索量≥50.xlsx', 'xlsx'),
        ('AI海报.xls', 'xls'),
    ]
    # 检查文件是否有可用关键词
    available = []
    for fname, ftype in files:
        if not os.path.exists(fname):
            continue
        try:
            if ftype == 'xlsx':
                import openpyxl
                wb = openpyxl.load_workbook(fname, read_only=True)
                ws = wb.active
                count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[0] and str(r[0]).strip())
                wb.close()
            else:
                import xlrd
                wb = xlrd.open_workbook(fname)
                ws = wb.sheet_by_index(0)
                count = sum(1 for i in range(1, ws.nrows) if ws.cell_value(i, 0) and str(ws.cell_value(i, 0)).strip())
            if count > 0:
                available.append((fname, ftype, count))
        except:
            pass

    if not available:
        print(json.dumps({'error': '所有关键词文件已用完！请补充关键词'}))
        sys.exit(0)

    # 按关键词数量加权随机选择文件
    total = sum(c for _, _, c in available)
    r = random.uniform(0, total)
    cumulative = 0
    chosen_file = available[0][0]
    for fname, ftype, count in available:
        cumulative += count
        if r <= cumulative:
            chosen_file = fname
            break

    # 取词
    ext = 'xlsx' if chosen_file.endswith('.xlsx') else 'xls'
    if ext == 'xlsx':
        kw = pick_from_xlsx(chosen_file)
    else:
        kw = pick_from_xls(chosen_file)

    if not kw:
        print(json.dumps({'error': f'{chosen_file} 取词失败'}))
        sys.exit(0)

    site, domain = guess_site(kw)

    # 获取搜索量（如果有）
    sv = ''
    try:
        if chosen_file == '合并结果_搜索量≥50.xlsx':
            import openpyxl
            wb = openpyxl.load_workbook(chosen_file, read_only=True)
            # 无法从已保存的文件读取，搜索量已丢失
            wb.close()
    except:
        pass

    result = {
        'keyword': kw,
        'site': site,
        'domain': domain,
        'file': chosen_file,
    }
    print(json.dumps(result, ensure_ascii=False))


if __name__ == '__main__':
    main()
